import openpyxl

# 읽어야 할 파일과 쓸 엑셀 파일 경로 설정
input_file = 'src_urr.txt'
output_file = 'ELG_GW-C_PCC.xlsx'

# 엑셀 파일 열기 (존재하지 않으면 새로 생성)
try:
    workbook = openpyxl.load_workbook(output_file)
except FileNotFoundError:
    workbook = openpyxl.Workbook()

# 첫 번째 시트 활성화
sheet = workbook['URR']
for row in range(2, sheet.max_row + 1):
    sheet.delete_rows(row, 1)

# 두 번째 행부터 데이터를 채우기 시작
row = 2

# 필요한 정보들을 담을 변수 초기화
AAAA = BBBB = CCCC = DDDD = ""
rg_flag = False
sid_flag = False

# 이전 AAAA와 BBBB 값을 저장할 변수
prev_AAAA = prev_BBBB = ""

with open(input_file, 'r') as file:
    for line in file:
        line = line.strip()  # 앞뒤 공백 제거

        if line.startswith("rating-group"):
            AAAA = line.split("rating-group", 1)[1].strip()
            rg_flag = True

        elif line.startswith("offline-urr-id") and rg_flag:
            BBBB = line.split("offline-urr-id", 1)[1].strip()
            rg_flag = False

        elif line.startswith("service-id"):
            CCCC = line.split("service-id", 1)[1].strip()
            sid_flag = True

        elif line.startswith("offline-urr-id") and sid_flag:
            DDDD = line.split("offline-urr-id", 1)[1].strip()
            sid_flag = False

        elif line.startswith("!"):  # 세트가 끝날 때 엑셀에 작성
            # CCCC 값이 없으면 데이터를 입력하지 않음
            if CCCC:
                # AAAA와 BBBB 값이 없을 경우 이전 값을 사용
                if not AAAA:
                    AAAA = prev_AAAA
                if not BBBB:
                    BBBB = prev_BBBB

                sheet[f'A{row}'] = AAAA
                sheet[f'B{row}'] = BBBB
                sheet[f'C{row}'] = CCCC
                sheet[f'D{row}'] = DDDD

                # 이전 값 업데이트
                prev_AAAA = AAAA
                prev_BBBB = BBBB

                # 다음 행으로 이동
                row += 1

            # 변수 초기화
            AAAA = BBBB = CCCC = DDDD = ""

# 엑셀 파일 저장
workbook.save(output_file)

