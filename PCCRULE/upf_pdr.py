import openpyxl

# 파일 경로 설정
output_file = 'ELG_GW-U_PCC.xlsx'
src_pdr_file = 'src_pdr.txt'

# Excel 파일 불러오기
workbook = openpyxl.load_workbook(output_file)
sheet = workbook['PDR']

# src_pdr.txt 파일 읽기
with open(src_pdr_file, 'r') as file:
    src_pdr_lines = file.readlines()

# PDR 데이터를 저장할 리스트
pdr_data_list = []

pdr = ""
pdr_enabled = ""
pdr_precedence = ""
pdr_application = ""
pdr_urr = ""

# src_pdr.txt 파일의 데이터를 읽어 PDR 세트로 나누기
for line in src_pdr_lines:
    line = line.strip()

    if line.startswith("user-plane"):
        # PDR 값 추출 (공백으로 분리하여 마지막 요소 추출)
        pdr = line.split()[-1]

    elif line.startswith("enabled"):
        pdr_enabled = line.split()[-1]

    elif line.startswith("precedence"):
        pdr_precedence = line.split()[-1]

    elif line.startswith("application"):
        pdr_application = line.split()[-1]

    elif line.startswith("predefined-urrs"):
        # predefined-urrs 값 추출
        pdr_urr = line.split("[", 1)[1].split("]", 1)[0]

    elif line.startswith("!"):
        # 데이터를 pdr_data_list에 추가
        pdr_data_list.append({
            "pdr": pdr,
            "pdr_enabled": pdr_enabled,
            "pdr_precedence": pdr_precedence,
            "pdr_application": pdr_application,
            "pdr_urr": pdr_urr
        })

        # 변수 초기화
        pdr = ""
        pdr_enabled = ""
        pdr_precedence = ""
        pdr_application = ""
        pdr_urr = ""

# 엑셀 파일에서 C컬럼과 비교하여 값을 입력하거나 새로운 행에 데이터를 추가
for pdr_data in pdr_data_list:
    pdr_value = pdr_data['pdr']
    row_found = False

    # C컬럼에서 pdr_value와 일치하는 값 찾기
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=3).value
        if cell_value == pdr_value:
            # 일치하는 경우 D, E, F, G 컬럼에 값 입력
            sheet.cell(row=row, column=4).value = pdr_data['pdr_enabled']
            sheet.cell(row=row, column=5).value = pdr_data['pdr_precedence']
            sheet.cell(row=row, column=6).value = pdr_data['pdr_application']
            sheet.cell(row=row, column=7).value = pdr_data['pdr_urr']
            row_found = True

    if not row_found:
        # 일치하는 값이 없으면 새로운 행에 데이터 추가
        new_row = sheet.max_row + 1
        sheet.cell(row=new_row, column=3).value = pdr_data['pdr']
        sheet.cell(row=new_row, column=4).value = pdr_data['pdr_enabled']
        sheet.cell(row=new_row, column=5).value = pdr_data['pdr_precedence']
        sheet.cell(row=new_row, column=6).value = pdr_data['pdr_application']
        sheet.cell(row=new_row, column=7).value = pdr_data['pdr_urr']

# 변경사항 저장
workbook.save(output_file)

