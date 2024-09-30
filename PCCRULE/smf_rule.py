import openpyxl

# 읽어야 할 파일과 쓸 엑셀 파일 경로 설정
input_file = 'src_rule.txt'
output_file = 'ELG_GW-C_PCC.xlsx'

# 엑셀 파일 열기 (존재하지 않으면 새로 생성)
try:
    workbook = openpyxl.load_workbook(output_file)
except FileNotFoundError:
    workbook = openpyxl.Workbook()

# 첫 번째 시트 활성화
sheet = workbook['PCC']
#for row in range(2, sheet.max_row + 1):
#        sheet.delete_rows(row, 1)
sheet.delete_rows(2, sheet.max_row)

# 두 번째 행부터 데이터를 채우기 시작
row = 2

# 필요한 정보들을 담을 변수 초기화
pcc_rule = description = rating_group = offline = online = ""
service_qos = gate_status = traffic_redirection_rule = user_plane_rule = ""

with open(input_file, 'r') as file:
    for line in file:
        line = line.strip()  # 앞뒤 공백 제거

        if line.startswith("pcc-rule"):
            pcc_rule = line.split("pcc-rule", 1)[1].strip()

        elif line.startswith("description"):
            description = line.split("description", 1)[1].strip()

        elif line.startswith("rating-group"):
            rating_group = line.split("rating-group", 1)[1].strip()

        elif line.startswith("offline"):
            offline = line.split("offline", 1)[1].strip()

        elif line.startswith("online"):
            online = line.split("online", 1)[1].strip()

        elif "service-qos" in line and "shared-profile" in line:
            service_qos = line.split("shared-profile", 1)[1].strip()

        elif line.startswith("gate-status"):
            gate_status = line.split("gate-status", 1)[1].strip()

        elif line.startswith("traffic-redirection-rule"):
            traffic_redirection_rule = line.split("traffic-redirection-rule", 1)[1].strip()

        elif line.startswith("user-plane-rule"):
            user_plane_rule = line.split("[", 1)[1].split("]", 1)[0].replace(" ", "")  # 공백 제거

        elif line.startswith("!"):  # 세트가 끝날 때 엑셀에 작성
            sheet[f'C{row}'] = pcc_rule
            sheet[f'D{row}'] = description
            sheet[f'E{row}'] = rating_group
            sheet[f'F{row}'] = offline
            sheet[f'G{row}'] = online
            sheet[f'H{row}'] = service_qos
            sheet[f'K{row}'] = gate_status
            sheet[f'L{row}'] = traffic_redirection_rule
            sheet[f'N{row}'] = user_plane_rule
            
            # 다음 행으로 이동
            row += 1

            # 변수 초기화
            pcc_rule = description = rating_group = offline = online = ""
            service_qos = gate_status = traffic_redirection_rule = user_plane_rule = ""

# 엑셀 파일 저장
workbook.save(output_file)

