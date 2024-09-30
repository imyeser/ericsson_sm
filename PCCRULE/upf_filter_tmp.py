import openpyxl

# src_filter.txt 파일 읽기
with open('src_filter.txt', 'r') as file:
    lines = file.readlines()

# output3.xlsx 파일 열기
workbook = openpyxl.load_workbook('ELG_GW-U_PCC.xlsx')
sheet = workbook['FILTER']
#for row in range(2, sheet.max_row + 1):
#    sheet.delete_rows(row, 1)
sheet.delete_rows(2, sheet.max_row)

filter = port = ue = addr = ""

filter_info = []
for line in lines:
    if 'filters' in line:
        filter = line.split()[-1]
    elif 'ip-flow-network-port' in line:
        port = line[line.index('[')+1:line.index(']')]
    elif 'ip-flow-ue-address' in line:
        ue = line[line.index('[')+1:line.index(']')]
    elif 'ip-flow-network-address' in line:
        addrs = line[line.index('[')+1:line.index(']')].split()
        for addr in addrs:
            sheet.append([filter, port, addr, ue])
    filter = port = ue = addr = ""
# 변경 내용을 저장
workbook.save('ELG_GW-U_PCC.xlsx')
