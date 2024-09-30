import openpyxl

# src_filter.txt 파일 읽기
with open('src_prule.txt', 'r') as file:
    lines = file.readlines()

workbook = openpyxl.load_workbook('ELG_GW-U_PCC.xlsx')
sheet = workbook['PDR']
#for row in range(2, sheet.max_row + 1):
#    sheet.delete_rows(row, 1)
sheet.delete_rows(2, sheet.max_row)

fil = port = ue = addr = ""

filter_info = []
for line in lines:
    if 'predefined-rules' in line:
        rule = line.split()[-1]
    elif 'enabled' in line:
        enabled = line.split()[-1]
    elif 'predefined-pdr' in line:
        pdrs = line[line.index('[')+1:line.index(']')].split()
        for pdr in pdrs:
            sheet.append([rule, enabled, pdr])
        rule = enabled = pdr = ""
# 변경 내용을 저장
workbook.save('ELG_GW-U_PCC.xlsx')
